from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import sqlite3, json, os, io, csv
from datetime import datetime

# ── ReportLab font setup (fixes rupee ■ black-square bug) ───────────────────
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase.pdfmetrics import registerFontFamily

_FONT_DIRS = [
    '/usr/share/fonts/truetype/liberation',
    '/usr/share/fonts/truetype/freefont',
    '/usr/share/fonts/truetype/dejavu',
    'C:/Windows/Fonts', '/Library/Fonts',
    os.path.expanduser('~/Library/Fonts'),
]
def _find_font(names):
    for d in _FONT_DIRS:
        for n in names:
            p = os.path.join(d, n)
            if os.path.exists(p): return p
    return None

_F   = _find_font(['LiberationSans-Regular.ttf',     'FreeSans.ttf',             'DejaVuSans.ttf'])
_FB  = _find_font(['LiberationSans-Bold.ttf',        'FreeSansBold.ttf',         'DejaVuSans-Bold.ttf'])
_FI  = _find_font(['LiberationSans-Italic.ttf',      'FreeSansOblique.ttf',      'DejaVuSans-Oblique.ttf'])
_FBI = _find_font(['LiberationSans-BoldItalic.ttf',  'FreeSansBoldOblique.ttf',  'DejaVuSans-Bold.ttf'])

if _F:
    pdfmetrics.registerFont(TTFont('AppFont',        _F))
    pdfmetrics.registerFont(TTFont('AppFont-Bold',   _FB or _F))
    pdfmetrics.registerFont(TTFont('AppFont-Italic', _FI or _F))
    pdfmetrics.registerFont(TTFont('AppFont-BI',     _FBI or _FB or _F))
    registerFontFamily('AppFont', normal='AppFont', bold='AppFont-Bold',
                       italic='AppFont-Italic', boldItalic='AppFont-BI')
    BASE_FONT = 'AppFont'; BASE_BOLD = 'AppFont-Bold'
else:
    BASE_FONT = 'Helvetica'; BASE_BOLD = 'Helvetica-Bold'

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from openpyxl import Workbook
from openpyxl.styles import Font as XFont, PatternFill, Alignment as XAlign, Border as XBorder, Side

try:
    from num2words import num2words as _n2w
    def _words(n):
        try:
            r = int(n); p = round((n - r) * 100)
            w = _n2w(r, lang='en_IN').title()
            out = f"Rupees {w}"
            if p: out += f" And {_n2w(p, lang='en_IN').title()} Paise"
            return out + " Only"
        except Exception: return ""
except ImportError:
    def _words(n): return ""

RUPEE = '\u20b9'  # ₹ — renders correctly via TTF; no black square

app = Flask(__name__)
CORS(app)
DB_PATH = os.environ.get('DB_PATH', 'gst_invoices.db')

# ─────────────────────────────────────────────────────────────────────────────
# DATABASE
# ─────────────────────────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn

def init_db():
    conn = get_db()
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS config (
            id INTEGER PRIMARY KEY,
            company_name TEXT, gstin TEXT, address TEXT,
            contact TEXT, email TEXT,
            bank_name TEXT, account_no TEXT, ifsc TEXT,
            branch TEXT, account_type TEXT, beneficiary_name TEXT,
            signatory_name TEXT, signatory_title TEXT,
            logo_url TEXT, state_code TEXT DEFAULT "27"
        );
        CREATE TABLE IF NOT EXISTS invoices (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            invoice_no     TEXT UNIQUE,
            invoice_date   TEXT,
            po_no          TEXT,
            client_name    TEXT,
            client_address TEXT,
            client_gstin   TEXT,
            supply_type    TEXT DEFAULT "intra",
            notes          TEXT,
            status         TEXT DEFAULT "draft",
            created_at     TEXT,
            updated_at     TEXT
        );
        CREATE TABLE IF NOT EXISTS line_items (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            invoice_id    INTEGER,
            sr_no         INTEGER,
            description   TEXT,
            hsn_sac       TEXT,
            qty           REAL,
            unit          TEXT,
            rate          REAL,
            taxable_value REAL,
            cgst_rate     REAL DEFAULT 0,
            cgst_amt      REAL DEFAULT 0,
            sgst_rate     REAL DEFAULT 0,
            sgst_amt      REAL DEFAULT 0,
            igst_rate     REAL DEFAULT 0,
            igst_amt      REAL DEFAULT 0,
            FOREIGN KEY(invoice_id) REFERENCES invoices(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS audit_log (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            ts        TEXT,
            action    TEXT,
            entity    TEXT,
            entity_id INTEGER,
            detail    TEXT
        );
    ''')
    conn.commit(); conn.close()

init_db()

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def log_action(action, entity, entity_id, detail=''):
    conn = get_db()
    conn.execute("INSERT INTO audit_log(ts,action,entity,entity_id,detail) VALUES(?,?,?,?,?)",
                 (datetime.now().isoformat(), action, entity, entity_id, str(detail)))
    conn.commit(); conn.close()

def _inv_totals(items):
    tv = sum(i['taxable_value'] or 0 for i in items)
    cg = sum(i['cgst_amt']      or 0 for i in items)
    sg = sum(i['sgst_amt']      or 0 for i in items)
    ig = sum(i['igst_amt']      or 0 for i in items)
    grand = tv + cg + sg + ig
    return dict(total_taxable=tv, total_cgst=cg, total_sgst=sg,
                total_igst=ig, total=round(grand), grand_exact=grand)

def _fetch_invoice(conn, inv_id):
    row = conn.execute('SELECT * FROM invoices WHERE id=?', (inv_id,)).fetchone()
    if not row: return None
    inv = dict(row)
    items = [dict(r) for r in conn.execute(
        'SELECT * FROM line_items WHERE invoice_id=? ORDER BY sr_no', (inv_id,))]
    inv['line_items'] = items
    inv.update(_inv_totals(items))
    return inv

def _upsert_invoice(conn, data, inv_id=None):
    now = datetime.now().isoformat()
    F = ('invoice_no','invoice_date','po_no','client_name',
         'client_address','client_gstin','supply_type','notes','status')
    vals = [data.get(f,'') for f in F]
    if inv_id:
        conn.execute(f"UPDATE invoices SET {','.join(f+'=?' for f in F)},updated_at=? WHERE id=?",
                     vals + [now, inv_id])
        conn.execute('DELETE FROM line_items WHERE invoice_id=?', (inv_id,))
    else:
        conn.execute(f"INSERT INTO invoices ({','.join(F)},created_at,updated_at)"
                     f" VALUES ({','.join('?'*len(F))},?,?)", vals + [now, now])
        inv_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
    for item in data.get('line_items', []):
        conn.execute('''INSERT INTO line_items
            (invoice_id,sr_no,description,hsn_sac,qty,unit,rate,taxable_value,
             cgst_rate,cgst_amt,sgst_rate,sgst_amt,igst_rate,igst_amt)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (inv_id, item['sr_no'], item['description'], item.get('hsn_sac',''),
             item['qty'], item.get('unit',''), item['rate'], item['taxable_value'],
             item.get('cgst_rate',0), item.get('cgst_amt',0),
             item.get('sgst_rate',0), item.get('sgst_amt',0),
             item.get('igst_rate',0), item.get('igst_amt',0)))
    return inv_id

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/config', methods=['GET'])
def get_config():
    conn = get_db()
    row = conn.execute('SELECT * FROM config WHERE id=1').fetchone()
    conn.close()
    return jsonify(dict(row) if row else {})

@app.route('/api/config', methods=['POST'])
def save_config():
    data = request.json
    fields = ['company_name','gstin','address','contact','email','bank_name',
              'account_no','ifsc','branch','account_type','beneficiary_name',
              'signatory_name','signatory_title','logo_url','state_code']
    conn = get_db()
    vals = [data.get(f,'') for f in fields]
    if conn.execute('SELECT id FROM config WHERE id=1').fetchone():
        conn.execute(f"UPDATE config SET {','.join(f+'=?' for f in fields)} WHERE id=1", vals)
    else:
        conn.execute(f"INSERT INTO config (id,{','.join(fields)}) VALUES (1,{','.join('?'*len(fields))})", vals)
    conn.commit(); conn.close()
    log_action('UPDATE','config',1)
    return jsonify({'status':'ok'})

# ─────────────────────────────────────────────────────────────────────────────
# INVOICES
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/invoices', methods=['GET'])
def list_invoices():
    conn = get_db()
    rows = conn.execute('SELECT * FROM invoices ORDER BY id DESC').fetchall()
    result = []
    for r in rows:
        inv = dict(r)
        items = [dict(i) for i in conn.execute(
            'SELECT * FROM line_items WHERE invoice_id=? ORDER BY sr_no', (r['id'],))]
        inv['line_items'] = items
        inv.update(_inv_totals(items))
        result.append(inv)
    conn.close()
    return jsonify(result)

@app.route('/api/invoices/<int:inv_id>', methods=['GET'])
def get_invoice(inv_id):
    conn = get_db()
    inv = _fetch_invoice(conn, inv_id)
    conn.close()
    if not inv: return jsonify({'error':'Not found'}), 404
    return jsonify(inv)

@app.route('/api/invoices', methods=['POST'])
def create_invoice():
    data = request.json
    conn = get_db()
    try:
        inv_id = _upsert_invoice(conn, data)
        conn.commit()
        log_action('CREATE','invoice',inv_id,data.get('invoice_no',''))
        return jsonify({'id':inv_id,'status':'created'})
    except Exception as e:
        conn.rollback(); return jsonify({'error':str(e)}), 400
    finally: conn.close()

@app.route('/api/invoices/<int:inv_id>', methods=['PUT'])
def update_invoice(inv_id):
    data = request.json
    conn = get_db()
    try:
        _upsert_invoice(conn, data, inv_id)
        conn.commit()
        log_action('UPDATE','invoice',inv_id,data.get('invoice_no',''))
        return jsonify({'status':'updated'})
    except Exception as e:
        conn.rollback(); return jsonify({'error':str(e)}), 400
    finally: conn.close()

@app.route('/api/invoices/<int:inv_id>', methods=['DELETE'])
def delete_invoice(inv_id):
    conn = get_db()
    row = conn.execute('SELECT invoice_no FROM invoices WHERE id=?', (inv_id,)).fetchone()
    conn.execute('DELETE FROM invoices WHERE id=?', (inv_id,))
    conn.commit(); conn.close()
    if row: log_action('DELETE','invoice',inv_id,row['invoice_no'])
    return jsonify({'status':'deleted'})

@app.route('/api/next-invoice-no', methods=['GET'])
def next_invoice_no():
    conn = get_db()
    cfg   = conn.execute('SELECT * FROM config WHERE id=1').fetchone()
    count = conn.execute('SELECT COUNT(*) as c FROM invoices').fetchone()['c']
    conn.close()
    prefix = ''.join(w[0].upper() for w in (cfg['company_name'] or '').split() if w)[:4] if cfg else 'INV'
    prefix = prefix or 'INV'
    now = datetime.now()
    fy  = f"{now.year}-{str(now.year+1)[2:]}" if now.month >= 4 else f"{now.year-1}-{str(now.year)[2:]}"
    return jsonify({'invoice_no': f"{prefix}/{count+1}/{fy}"})

# ─────────────────────────────────────────────────────────────────────────────
# PDF EXPORT
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/invoices/<int:inv_id>/pdf', methods=['GET'])
def export_pdf(inv_id):
    conn = get_db()
    inv  = _fetch_invoice(conn, inv_id)
    cfg  = conn.execute('SELECT * FROM config WHERE id=1').fetchone()
    conn.close()
    if not inv: return jsonify({'error':'Not found'}), 404
    cfg = dict(cfg) if cfg else {}

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             topMargin=8*mm, bottomMargin=10*mm,
                             leftMargin=10*mm, rightMargin=10*mm)
    W = 190*mm
    hdr_blue = colors.HexColor('#BDD7EE')
    grn_bg   = colors.HexColor('#E2EFDA')
    navy     = colors.HexColor('#1F4E79')
    story    = []
    THIN     = 0.4

    def P(text, bold=False, size=8.5, align=TA_LEFT, color=colors.black):
        fn = BASE_BOLD if bold else BASE_FONT
        st = ParagraphStyle('_', fontName=fn, fontSize=size,
                             textColor=color, alignment=align, leading=size*1.4)
        return Paragraph(str(text or ''), st)

    def money(v): return f"{RUPEE}{v:,.2f}"

    # Header
    t = Table([[P('TAX INVOICE', bold=True, size=15, align=TA_CENTER, color=navy)],
               [P(cfg.get('company_name',''), bold=True, size=11, align=TA_CENTER)],
               [P(f"GSTIN: {cfg.get('gstin','')}  \u2022  Ph: {cfg.get('contact','')}  \u2022  {cfg.get('email','')}", size=7.5, align=TA_CENTER)],
               [P(cfg.get('address',''), size=7.5, align=TA_CENTER)]],
              colWidths=[W])
    t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),hdr_blue),
                            ('BOX',(0,0),(-1,-1),THIN,colors.grey),
                            ('TOPPADDING',(0,0),(-1,-1),4),
                            ('BOTTOMPADDING',(0,0),(-1,-1),4)]))
    story.append(t); story.append(Spacer(1,2*mm))

    # Meta
    t = Table([[P('Invoice No:',bold=True), P(inv['invoice_no']),
                P('Date:',bold=True),       P(inv['invoice_date'])],
               [P('PO No:',bold=True),      P(inv.get('po_no','')),
                P('Status:',bold=True),     P((inv.get('status') or '').upper())]],
              colWidths=[W*.14,W*.36,W*.12,W*.38])
    t.setStyle(TableStyle([('BOX',(0,0),(-1,-1),THIN,colors.grey),
                            ('INNERGRID',(0,0),(-1,-1),.2,colors.lightgrey),
                            ('TOPPADDING',(0,0),(-1,-1),3),
                            ('BOTTOMPADDING',(0,0),(-1,-1),3)]))
    story.append(t); story.append(Spacer(1,2*mm))

    # Bill To
    t = Table([[P('BILL TO',bold=True,color=navy,size=8)],
               [P(inv.get('client_name',''),bold=True,size=10)],
               [P(inv.get('client_address',''),size=7.5)],
               [P(f"GSTIN: {inv.get('client_gstin','')}",size=8)]],
              colWidths=[W])
    t.setStyle(TableStyle([('BACKGROUND',(0,0),(0,0),colors.HexColor('#E9EFF7')),
                            ('BOX',(0,0),(-1,-1),THIN,colors.grey),
                            ('TOPPADDING',(0,0),(-1,-1),3),
                            ('BOTTOMPADDING',(0,0),(-1,-1),3)]))
    story.append(t); story.append(Spacer(1,3*mm))

    # Line items
    supply = inv.get('supply_type','intra')
    items  = inv['line_items']
    if supply == 'intra':
        cols = ['Sr','Description','HSN/SAC','Qty','Unit','Rate',
                'Taxable Value','CGST%','CGST Amt','SGST%','SGST Amt','Total']
        cw   = [W*.04,W*.21,W*.08,W*.06,W*.05,W*.08,
                W*.09,W*.05,W*.08,W*.05,W*.08,W*.09]
    else:
        cols = ['Sr','Description','HSN/SAC','Qty','Unit','Rate',
                'Taxable Value','IGST%','IGST Amt','Total']
        cw   = [W*.04,W*.26,W*.09,W*.07,W*.06,W*.09,W*.12,W*.06,W*.12,W*.09]

    rows = [[P(c,bold=True,align=TA_CENTER,size=7.5) for c in cols]]
    tv_tot=cg_tot=sg_tot=ig_tot=0

    for item in items:
        tv=item['taxable_value'] or 0; ca=item['cgst_amt'] or 0
        sa=item['sgst_amt'] or 0;     ia=item['igst_amt'] or 0
        lt=tv+ca+sa+ia
        tv_tot+=tv; cg_tot+=ca; sg_tot+=sa; ig_tot+=ia
        if supply=='intra':
            row=[P(str(item['sr_no']),align=TA_CENTER),
                 P(item['description']),
                 P(str(item.get('hsn_sac','')),align=TA_CENTER),
                 P(f"{item['qty']:,.2f}",align=TA_RIGHT),
                 P(item.get('unit',''),align=TA_CENTER),
                 P(money(item['rate']),align=TA_RIGHT),
                 P(money(tv),align=TA_RIGHT),
                 P(f"{item['cgst_rate']}%",align=TA_CENTER),
                 P(money(ca),align=TA_RIGHT),
                 P(f"{item['sgst_rate']}%",align=TA_CENTER),
                 P(money(sa),align=TA_RIGHT),
                 P(money(lt),align=TA_RIGHT,bold=True)]
        else:
            row=[P(str(item['sr_no']),align=TA_CENTER),
                 P(item['description']),
                 P(str(item.get('hsn_sac','')),align=TA_CENTER),
                 P(f"{item['qty']:,.2f}",align=TA_RIGHT),
                 P(item.get('unit',''),align=TA_CENTER),
                 P(money(item['rate']),align=TA_RIGHT),
                 P(money(tv),align=TA_RIGHT),
                 P(f"{item['igst_rate']}%",align=TA_CENTER),
                 P(money(ia),align=TA_RIGHT),
                 P(money(lt),align=TA_RIGHT,bold=True)]
        rows.append(row)

    grand=tv_tot+cg_tot+sg_tot+ig_tot; rounded=round(grand); roff=rounded-grand
    if supply=='intra':
        tr=[P(''),P('TOTAL',bold=True),P(''),P(''),P(''),P(''),
            P(money(tv_tot),bold=True,align=TA_RIGHT),P(''),
            P(money(cg_tot),bold=True,align=TA_RIGHT),P(''),
            P(money(sg_tot),bold=True,align=TA_RIGHT),
            P(money(grand), bold=True,align=TA_RIGHT)]
    else:
        tr=[P(''),P('TOTAL',bold=True),P(''),P(''),P(''),P(''),
            P(money(tv_tot),bold=True,align=TA_RIGHT),P(''),
            P(money(ig_tot),bold=True,align=TA_RIGHT),
            P(money(grand), bold=True,align=TA_RIGHT)]
    rows.append(tr)

    t=Table(rows,colWidths=cw,repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),hdr_blue),
        ('BACKGROUND',(0,-1),(-1,-1),grn_bg),
        ('BOX',(0,0),(-1,-1),THIN,colors.grey),
        ('INNERGRID',(0,0),(-1,-1),.2,colors.HexColor('#D0D0D0')),
        ('FONTSIZE',(0,0),(-1,-1),8),
        ('TOPPADDING',(0,0),(-1,-1),2),
        ('BOTTOMPADDING',(0,0),(-1,-1),2),
        ('ROWBACKGROUNDS',(0,1),(-1,-2),[colors.white,colors.HexColor('#F7FBFF')]),
    ]))
    story.append(t); story.append(Spacer(1,2*mm))

    # Summary
    ts=Table([[P('Round Off',align=TA_RIGHT),     P(f"{RUPEE}{roff:+.2f}",align=TA_RIGHT)],
              [P('TOTAL INVOICE VALUE',bold=True,size=10,align=TA_RIGHT),
               P(money(rounded),bold=True,size=10,align=TA_RIGHT)],
              [P('Amount in Words:',bold=True),  P(_words(rounded),size=7.5)]],
             colWidths=[W*.6,W*.4])
    ts.setStyle(TableStyle([('BOX',(0,0),(-1,-1),THIN,colors.grey),
                             ('INNERGRID',(0,0),(-1,-1),.2,colors.lightgrey),
                             ('BACKGROUND',(0,1),(-1,1),grn_bg),
                             ('TOPPADDING',(0,0),(-1,-1),3),
                             ('BOTTOMPADDING',(0,0),(-1,-1),3)]))
    story.append(ts); story.append(Spacer(1,3*mm))

    # GST Breakup
    gst_rows=[[P('HSN/SAC',bold=True,align=TA_CENTER),
               P('Taxable Value',bold=True,align=TA_RIGHT),
               P('CGST',bold=True,align=TA_RIGHT),
               P('SGST',bold=True,align=TA_RIGHT),
               P('IGST',bold=True,align=TA_RIGHT),
               P('Total Tax',bold=True,align=TA_RIGHT)]]
    for item in items:
        tv=item['taxable_value'] or 0; ca=item['cgst_amt'] or 0
        sa=item['sgst_amt'] or 0;     ia=item['igst_amt'] or 0
        gst_rows.append([P(str(item.get('hsn_sac','')),align=TA_CENTER),
                          P(money(tv),align=TA_RIGHT), P(money(ca),align=TA_RIGHT),
                          P(money(sa),align=TA_RIGHT), P(money(ia),align=TA_RIGHT),
                          P(money(ca+sa+ia),bold=True,align=TA_RIGHT)])
    gst_rows.append([P('Total',bold=True),
                     P(money(tv_tot),bold=True,align=TA_RIGHT),
                     P(money(cg_tot),bold=True,align=TA_RIGHT),
                     P(money(sg_tot),bold=True,align=TA_RIGHT),
                     P(money(ig_tot),bold=True,align=TA_RIGHT),
                     P(money(cg_tot+sg_tot+ig_tot),bold=True,align=TA_RIGHT)])
    tg=Table(gst_rows,colWidths=[W*.14,W*.18,W*.16,W*.16,W*.16,W*.20])
    tg.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#F0F4F8')),
                             ('BACKGROUND',(0,-1),(-1,-1),grn_bg),
                             ('BOX',(0,0),(-1,-1),THIN,colors.grey),
                             ('INNERGRID',(0,0),(-1,-1),.2,colors.lightgrey),
                             ('FONTSIZE',(0,0),(-1,-1),8),
                             ('TOPPADDING',(0,0),(-1,-1),2),
                             ('BOTTOMPADDING',(0,0),(-1,-1),2)]))
    story.append(P('GST Breakup',bold=True,size=8.5))
    story.append(Spacer(1,1*mm)); story.append(tg); story.append(Spacer(1,3*mm))

    # Bank + Signature
    bank_text=(f"<b>Bank:</b> {cfg.get('bank_name','')}<br/>"
               f"<b>Account No:</b> {cfg.get('account_no','')}<br/>"
               f"<b>IFSC:</b> {cfg.get('ifsc','')} &nbsp; <b>Branch:</b> {cfg.get('branch','')}<br/>"
               f"<b>A/C Type:</b> {cfg.get('account_type','')}<br/>"
               f"<b>Beneficiary:</b> {cfg.get('beneficiary_name','')}")
    sig_text=(f"For {cfg.get('company_name','')}"
              f"<br/><br/><br/><br/>"
              f"<b>{cfg.get('signatory_name','')}</b><br/>"
              f"{cfg.get('signatory_title','')}")
    note="This is a computer-generated invoice."
    if inv.get('notes'): note+=f"<br/><b>Notes:</b> {inv['notes']}"
    bot=Table([[P(bank_text,size=8),P(sig_text,size=8,align=TA_CENTER)],
               [P(note,size=7,color=colors.grey),P('')]],
              colWidths=[W*.6,W*.4])
    bot.setStyle(TableStyle([('BOX',(0,0),(-1,-1),THIN,colors.grey),
                              ('INNERGRID',(0,0),(-1,-1),.2,colors.lightgrey),
                              ('TOPPADDING',(0,0),(-1,-1),5),
                              ('BOTTOMPADDING',(0,0),(-1,-1),8)]))
    story.append(bot)
    doc.build(story)
    buf.seek(0)
    safe_no = (inv['invoice_no'] or 'invoice').replace('/','_')
    return send_file(buf, mimetype='application/pdf',
                     download_name=f"Invoice_{safe_no}.pdf")

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/invoices/<int:inv_id>/excel', methods=['GET'])
def export_excel(inv_id):
    conn = get_db()
    inv  = _fetch_invoice(conn, inv_id)
    cfg  = conn.execute('SELECT * FROM config WHERE id=1').fetchone()
    conn.close()
    if not inv: return jsonify({'error':'Not found'}), 404
    cfg = dict(cfg) if cfg else {}

    wb=Workbook(); ws=wb.active; ws.title="Invoice"
    blue_fill=PatternFill("solid",fgColor="9CC2E5")
    yel_fill =PatternFill("solid",fgColor="FFFF00")
    grn_fill =PatternFill("solid",fgColor="92D050")
    hfont=XFont(name='Arial',bold=True,size=9)
    nfont=XFont(name='Arial',size=9)
    ctr=XAlign(horizontal='center',vertical='center',wrap_text=True)
    rgt=XAlign(horizontal='right', vertical='center')
    lft=XAlign(horizontal='left',  vertical='center',wrap_text=True)
    brd=XBorder(left=Side(style='thin'),right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

    def sc(cell,fill=None,font=None,align=None,fmt=None):
        if fill:  cell.fill=fill
        if font:  cell.font=font
        if align: cell.alignment=align
        cell.border=brd
        if fmt:   cell.number_format=fmt

    ws.merge_cells('A1:P1'); ws['A1']='TAX INVOICE'
    sc(ws['A1'],blue_fill,XFont(name='Arial',bold=True,size=16),ctr)
    ws.merge_cells('A2:G2'); ws['A2']=cfg.get('company_name','')
    sc(ws['A2'],blue_fill,XFont(name='Arial',bold=True,size=12),ctr)
    ws['H2']='Invoice No:'; sc(ws['H2'],blue_fill,hfont,lft)
    ws.merge_cells('K2:P2'); ws['K2']=inv['invoice_no']
    sc(ws['K2'],yel_fill,nfont,lft)
    ws['H3']='Date:'; sc(ws['H3'],blue_fill,hfont)
    ws.merge_cells('K3:P3'); ws['K3']=inv['invoice_date']
    sc(ws['K3'],yel_fill,nfont)
    ws['H4']='PO No:'; sc(ws['H4'],blue_fill,hfont)
    ws.merge_cells('K4:P4'); ws['K4']=inv.get('po_no','')
    sc(ws['K4'],yel_fill,nfont)
    ws.merge_cells('A3:G4')
    ws['A3']=f"GSTIN: {cfg.get('gstin','')} | Ph: {cfg.get('contact','')} | {cfg.get('email','')}"
    sc(ws['A3'],blue_fill,nfont,lft)
    ws.merge_cells('A5:P5'); ws['A5']=cfg.get('address','')
    sc(ws['A5'],blue_fill,nfont,lft)
    ws.merge_cells('A6:G6'); ws['A6']='BILLING DETAILS'
    sc(ws['A6'],blue_fill,hfont,ctr)
    ws.merge_cells('A7:G7'); ws['A7']=inv['client_name']
    sc(ws['A7'],yel_fill,XFont(name='Arial',bold=True,size=10),lft)
    ws.merge_cells('A8:G9'); ws['A8']=inv.get('client_address','')
    sc(ws['A8'],yel_fill,nfont,lft)
    ws.merge_cells('A10:G10'); ws['A10']=f"GSTIN: {inv.get('client_gstin','')}"
    sc(ws['A10'],yel_fill,nfont)

    for col,lbl in [('A','SR'),('C','DESCRIPTION'),('F','HSN/SAC'),
                    ('G','QTY'),('H','UNIT'),('I','RATE'),('J','TAXABLE VALUE'),
                    ('K','CGST%'),('L','CGST AMT'),('M','SGST%'),('N','SGST AMT'),
                    ('O','IGST%'),('P','IGST AMT')]:
        ws[f'{col}12']=lbl; sc(ws[f'{col}12'],blue_fill,hfont,ctr)

    items=inv['line_items']; ds=13
    for idx,item in enumerate(items):
        r=ds+idx
        ws[f'A{r}']=item['sr_no']; sc(ws[f'A{r}'],yel_fill,nfont,ctr)
        ws.merge_cells(f'C{r}:E{r}'); ws[f'C{r}']=item['description']
        sc(ws[f'C{r}'],yel_fill,nfont,lft)
        ws[f'F{r}']=item.get('hsn_sac',''); sc(ws[f'F{r}'],yel_fill,nfont,ctr)
        ws[f'G{r}']=item['qty'];  sc(ws[f'G{r}'],yel_fill,nfont,rgt,'#,##0.00')
        ws[f'H{r}']=item.get('unit',''); sc(ws[f'H{r}'],yel_fill,nfont,ctr)
        ws[f'I{r}']=item['rate']; sc(ws[f'I{r}'],yel_fill,nfont,rgt,'#,##0.00')
        ws[f'J{r}']=item['taxable_value']; sc(ws[f'J{r}'],grn_fill,nfont,rgt,'#,##0.00')
        ws[f'K{r}']=item.get('cgst_rate',0); sc(ws[f'K{r}'],yel_fill,nfont,ctr)
        ws[f'L{r}']=item.get('cgst_amt',0); sc(ws[f'L{r}'],grn_fill,nfont,rgt,'#,##0.00')
        ws[f'M{r}']=item.get('sgst_rate',0); sc(ws[f'M{r}'],yel_fill,nfont,ctr)
        ws[f'N{r}']=item.get('sgst_amt',0); sc(ws[f'N{r}'],grn_fill,nfont,rgt,'#,##0.00')
        ws[f'O{r}']=item.get('igst_rate',0); sc(ws[f'O{r}'],yel_fill,nfont,ctr)
        ws[f'P{r}']=item.get('igst_amt',0); sc(ws[f'P{r}'],grn_fill,nfont,rgt,'#,##0.00')

    tot=_inv_totals(items); tr=ds+len(items)
    ws.merge_cells(f'A{tr}:I{tr}'); ws[f'A{tr}']='SUB-TOTAL'
    sc(ws[f'A{tr}'],blue_fill,hfont,ctr)
    for col,key in [('J','total_taxable'),('L','total_cgst'),('N','total_sgst'),('P','total_igst')]:
        ws[f'{col}{tr}']=tot[key]; sc(ws[f'{col}{tr}'],grn_fill,hfont,rgt,'#,##0.00')
    tr+=1; ws.merge_cells(f'A{tr}:I{tr}'); ws[f'A{tr}']='Round Off'
    sc(ws[f'A{tr}'],blue_fill,nfont)
    ws[f'J{tr}']=round(tot['grand_exact'])-tot['grand_exact']
    sc(ws[f'J{tr}'],grn_fill,nfont,rgt,'+#,##0.00;-#,##0.00;"-"')
    tr+=1; ws.merge_cells(f'A{tr}:I{tr}'); ws[f'A{tr}']='TOTAL INVOICE VALUE'
    sc(ws[f'A{tr}'],blue_fill,hfont)
    ws[f'J{tr}']=tot['total']
    sc(ws[f'J{tr}'],grn_fill,XFont(name='Arial',bold=True,size=11),rgt,'#,##0')
    tr+=1; ws.merge_cells(f'A{tr}:I{tr}'); ws[f'A{tr}']='Amount in Words'
    sc(ws[f'A{tr}'],blue_fill,hfont)
    ws.merge_cells(f'J{tr}:P{tr}'); ws[f'J{tr}']=_words(tot['total'])
    sc(ws[f'J{tr}'],grn_fill,nfont,lft)

    tr+=2
    for label,val in [('Bank',cfg.get('bank_name','')),
                       ('Account No',cfg.get('account_no','')),
                       ('IFSC',cfg.get('ifsc','')),
                       ('Branch',cfg.get('branch','')),
                       ('A/C Type',cfg.get('account_type','')),
                       ('Beneficiary',cfg.get('beneficiary_name',''))]:
        ws[f'A{tr}']=label; sc(ws[f'A{tr}'],blue_fill,hfont)
        ws.merge_cells(f'C{tr}:H{tr}'); ws[f'C{tr}']=val
        sc(ws[f'C{tr}'],blue_fill,nfont); tr+=1

    for col,w in {'A':5,'B':2,'C':24,'D':4,'E':4,'F':10,'G':9,'H':7,'I':10,
                  'J':12,'K':7,'L':11,'M':7,'N':11,'O':7,'P':11}.items():
        ws.column_dimensions[col].width=w

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    safe_no=(inv['invoice_no'] or 'invoice').replace('/','_')
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        download_name=f"Invoice_{safe_no}.xlsx")

# ─────────────────────────────────────────────────────────────────────────────
# BACKUP / RESTORE / CSV  (production data safety)
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/backup', methods=['GET'])
def backup():
    conn = get_db()
    cfg  = conn.execute('SELECT * FROM config WHERE id=1').fetchone()
    invs = conn.execute('SELECT * FROM invoices ORDER BY id').fetchall()
    data = {'exported_at': datetime.now().isoformat(), 'version': '1.0',
            'config': dict(cfg) if cfg else {}, 'invoices': []}
    for r in invs:
        inv = dict(r)
        inv['line_items'] = [dict(i) for i in conn.execute(
            'SELECT * FROM line_items WHERE invoice_id=? ORDER BY sr_no', (r['id'],))]
        data['invoices'].append(inv)
    conn.close()
    buf  = io.BytesIO(json.dumps(data, indent=2, default=str).encode('utf-8'))
    fname= f"gst_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    buf.seek(0)
    return send_file(buf, mimetype='application/json', download_name=fname)

@app.route('/api/restore', methods=['POST'])
def restore():
    # Support both FormData (legacy) and raw JSON body
    if request.is_json:
        data = request.get_json(silent=True) or {}
    elif 'file' in request.files:
        try:
            data = json.loads(request.files['file'].read().decode('utf-8'))
        except Exception as e:
            return jsonify({'error': f'Invalid JSON: {e}'}), 400
    else:
        return jsonify({'error': 'No data provided'}), 400

    return _do_restore(data)

@app.route('/api/restore-json', methods=['POST'])
def restore_json():
    """Accept restore data as raw JSON body — works in all browsers without FormData."""
    data = request.get_json(silent=True)
    if not data:
        return jsonify({'error': 'Invalid or empty JSON body'}), 400
    return _do_restore(data)

def _do_restore(data):
    conn = get_db()
    skipped = created = 0
    try:
        cfg = data.get('config', {})
        if cfg and not conn.execute('SELECT id FROM config WHERE id=1').fetchone():
            fields = ['company_name','gstin','address','contact','email','bank_name',
                      'account_no','ifsc','branch','account_type','beneficiary_name',
                      'signatory_name','signatory_title','logo_url','state_code']
            conn.execute(f"INSERT OR IGNORE INTO config (id,{','.join(fields)}) VALUES (1,{','.join('?'*len(fields))})",
                         [cfg.get(f,'') for f in fields])
        for inv in data.get('invoices', []):
            if conn.execute('SELECT id FROM invoices WHERE invoice_no=?',(inv['invoice_no'],)).fetchone():
                skipped += 1; continue
            _upsert_invoice(conn, inv); created += 1
        conn.commit()
        log_action('RESTORE','backup',0,f"created={created} skipped={skipped}")
        return jsonify({'status':'ok','created':created,'skipped':skipped})
    except Exception as e:
        conn.rollback(); return jsonify({'error':str(e)}), 500
    finally: conn.close()

@app.route('/api/export-csv', methods=['GET'])
def export_csv():
    conn = get_db()
    invs = [dict(r) for r in conn.execute('SELECT * FROM invoices ORDER BY id').fetchall()]
    buf  = io.StringIO()
    w    = csv.writer(buf)
    w.writerow(['Invoice No','Date','PO No','Client','Client GSTIN',
                'Supply Type','Taxable Value','CGST','SGST','IGST','Total','Status'])
    for inv in invs:
        items = [dict(i) for i in conn.execute(
            'SELECT * FROM line_items WHERE invoice_id=?', (inv['id'],))]
        t = _inv_totals(items)
        w.writerow([inv['invoice_no'],inv['invoice_date'],inv['po_no'],
                    inv['client_name'],inv['client_gstin'],inv['supply_type'],
                    t['total_taxable'],t['total_cgst'],t['total_sgst'],
                    t['total_igst'],t['total'],inv['status']])
    conn.close()
    buf.seek(0)
    return send_file(io.BytesIO(buf.getvalue().encode('utf-8-sig')),
                     mimetype='text/csv',
                     download_name=f"invoices_{datetime.now().strftime('%Y%m%d')}.csv")

# ─────────────────────────────────────────────────────────────────────────────
# AUDIT LOG
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/audit-log', methods=['GET'])
def get_audit_log():
    conn = get_db()
    rows = conn.execute('SELECT * FROM audit_log ORDER BY id DESC LIMIT 200').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

# ─────────────────────────────────────────────────────────────────────────────
# REPORTING
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/reports/summary', methods=['GET'])
def report_summary():
    conn  = get_db()
    invs  = conn.execute('SELECT * FROM invoices ORDER BY invoice_date DESC').fetchall()
    items = conn.execute('SELECT * FROM line_items').fetchall()
    conn.close()
    item_map = {}
    for item in items:
        iid = item['invoice_id']
        if iid not in item_map: item_map[iid]={'taxable':0,'cgst':0,'sgst':0,'igst':0}
        item_map[iid]['taxable'] += item['taxable_value'] or 0
        item_map[iid]['cgst']    += item['cgst_amt']      or 0
        item_map[iid]['sgst']    += item['sgst_amt']      or 0
        item_map[iid]['igst']    += item['igst_amt']      or 0
    monthly={}
    for inv in invs:
        date=(inv['invoice_date'] or '')[:7] or 'Unknown'
        if date not in monthly: monthly[date]={'count':0,'total':0}
        monthly[date]['count']+=1
        t=item_map.get(inv['id'],{})
        monthly[date]['total']+=sum(t.values())
    tv=sum(v['taxable'] for v in item_map.values())
    cg=sum(v['cgst']    for v in item_map.values())
    sg=sum(v['sgst']    for v in item_map.values())
    ig=sum(v['igst']    for v in item_map.values())
    clients={}
    for inv in invs:
        n=inv['client_name'] or 'Unknown'
        if n not in clients: clients[n]={'count':0,'total':0}
        clients[n]['count']+=1
        t=item_map.get(inv['id'],{})
        clients[n]['total']+=sum(t.values())
    top5=sorted([{'name':k,'count':v['count'],'total':v['total']}
                 for k,v in clients.items()],key=lambda x:-x['total'])[:5]
    return jsonify({'total_invoices':len(invs),'total_taxable':tv,
                    'total_cgst':cg,'total_sgst':sg,'total_igst':ig,
                    'total_value':tv+cg+sg+ig,'monthly':monthly,'top_clients':top5})

# ─────────────────────────────────────────────────────────────────────────────
# HEALTH CHECK
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/health', methods=['GET'])
def health():
    try:
        conn=get_db(); conn.execute('SELECT 1'); conn.close()
        return jsonify({'status':'ok','db':'connected','ts':datetime.now().isoformat()})
    except Exception as e:
        return jsonify({'status':'error','message':str(e)}), 500

if __name__ == '__main__':
    port  = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('DEBUG','true').lower() == 'true'
    app.run(debug=debug, host='0.0.0.0', port=port)
